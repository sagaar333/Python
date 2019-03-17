import openpyxl

def xlsx_to_vcf():

    wb=openpyxl.load_workbook('names_input.xlsx')
    ws=wb.active
    fd=open('Vcard.vcf','w')
    str_start="BEGIN : VCARD" +"\n"+"VERSION:2.1"+"\n"
    str_end="END:VCARD"
    #fd.write(str1)
    
    
    for i in range(1,ws.max_row +1):
        fn=ws.cell(row=i,column=1).value
        org=ws.cell(row=i,column=4).value
        email=ws.cell(row=i,column=2).value
        phone=ws.cell(row=i,column=3).value
        fd.write(str_start)
        fd.write("FN:"+fn+"\n")
        fd.write("ORG:"+org+"\n")
        fd.write("TEL:"+str(phone)+"\n")
        fd.write("email:"+email+"\n")
        fd.write(str_end+"\n")

    fd.close()
    wb.close()
        
                    
        
        



def main():
    xlsx_to_vcf()

if __name__=='__main__':
    main()
